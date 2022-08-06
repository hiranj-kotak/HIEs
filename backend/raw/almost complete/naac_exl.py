from openpyxl import load_workbook
def naac_1():
    file_path = "../Institutions_accredited_by_NAAC_whose_accreditation_period_is_valid_08_07_22.xlsx"
    wb = load_workbook(file_path)
    ws = wb["Universities"]
    coloum = ws['C']
    data = []
    for  x in range(len(coloum)):
        a = coloum[x].value
        data.append(a)
    coloum4 = ws['F']
    data1 = []
    for  y in range(len(coloum4)):
        b = coloum4[y].value
        data1.append(b)

    coloum5 = ws['G']
    data2 = []
    for y in range(len(coloum4)):
        b = str(coloum5[y].value)
        data2.append(b)

    for x in range(6):
        del data[0]
        del  data1[0]
        del data2[0]
    data_new = []
    for item in data:
        # a = item
        # list = a.split(",")
        # list1 = list[0].split("[")
        # list2 = list1[0].split("(")
        # list3 = list2[0].split("\n")
        data_new.append(item)
    # d = dict(zip(data_new, data1))
    # print(d)
    # return d
    naac_grades =[]
    for i in range(len(data_new)):
        d1 = {"__id": i + 1, "instituteName": data_new[i], "NAAC_grade": data1[i], "valid  upto": data2[i]}

        naac_grades.append(d1)

        # naac_aa = {
        #     "data":naac_grades
        # }

    return naac_grades

def naac_2():
    file_path = "../Institutions_accredited_by_NAAC_whose_accreditation_period_is_valid_08_07_22.xlsx"
    wb = load_workbook(file_path)
    ws = wb["Colleges"]
    coloum = ws['B']
    data = []
    for  x in range(len(coloum)):
        a = coloum[x].value
        data.append(a)
    coloum4 = ws['E']
    data1 = []
    for  y in range(len(coloum4)):
        b = coloum4[y].value
        data1.append(b)

    coloum5 = ws['F']
    data2 = []
    for y in range(len(coloum4)):
        b = str(coloum5[y].value)
        data2.append(b)

    del data[0]
    del  data1[0]
    del data2[0]
    data_new = []
    for item in data:
        # a = item
        # list = a.split(",")
        # list1 = list[0].split("[")
        # list2 = list1[0].split("(")
        # list3 = list2[0].split("\n")
        # list4 = list3[0].split("–")
        data_new.append(item)
    # d = dict(zip(data_new, data1))
    # print(d)
    # return d
    naac_grades =[]
    for i in range(len(data_new)):
        d1 = {
            "__id": "",
            "instituteName":"Carusat",
            "NAAC_grade":"A++",
            "valid  upto":"today"
        }

        d1["instituteName"] = data_new[i]
        d1["__id"] = i+1
        d1["NAAC_grade"] = data1[i]
        d1["valid  upto"] = data2[i]
        naac_grades.append(d1)

        # naac_aa = {
        #     "data":naac_grades
        # }

    return naac_grades
def naac_3():
    file_path = "../Institutions_accredited_by_NAAC_whose_accreditation_period_is_valid_08_07_22.xlsx"
    wb = load_workbook(file_path)
    ws = wb["Transition Autonomous Colleges"]
    coloum = ws['C']
    data = []
    for  x in range(len(coloum)):
        a = coloum[x].value
        data.append(a)
    coloum4 = ws['F']
    data1 = []
    for  y in range(len(coloum4)):
        b = coloum4[y].value
        data1.append(b)

    coloum5 = ws['G']
    data2 = []
    for y in range(len(coloum4)):
        b = str(coloum5[y].value)
        data2.append(b)

    for x in range(6):
        del data[0]
        del  data1[0]
        del data2[0]
    data_new = []
    for item in data:
        a = item
        # list = a.split(",")
        # list1 = list[0].split("[")
        # list2 = list1[0].split("(")
        # list3 = list2[0].split("\n")
        data_new.append(item)
    # d = dict(zip(data_new, data1))
    # print(d)
    # return d
    naac_grades =[]
    for i in range(len(data_new)):
        d1 = {
            "__id": "",
            "instituteName":"Carusat",
            "NAAC_grade":"A++",
            "valid  upto":"today"
        }

        d1["instituteName"] = data_new[i]
        d1["__id"] = i+1
        d1["NAAC_grade"] = data1[i]
        d1["valid  upto"] = data2[i]
        naac_grades.append(d1)

        # naac_aa = {
        #     "data":naac_grades
        # }

    return naac_grades

# print(naac1())
def naac():
    hit = []
    hit+=naac_1()
    hit+=naac_2()
    hit+=naac_3()
    # print(hit)
    return hit

# with open("naac.txt", "w") as external_file:
#     print(naac(), file=external_file)
#     external_file.close()

'''don't call function for getting data always collect data from their .TXT file'''
