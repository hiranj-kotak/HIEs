import spacy
from openpyxl import load_workbook
import json
# from testcheck import naac
# institute_name = input()
nlp = spacy.load('en_core_web_sm')
file_path = "Institutions_Accredited_by_NAAC_Whose_Accreditation_Period_Is_Valid_uploaded_1_8_2022.xlsx"
sheet = ['Universities','Colleges','Transition Autonomous Colleges']
wb = load_workbook(file_path)
naac1 = []
for s in sheet:
    ws = wb[s]
    data = []
    # name = []
    coloum1 = ws['C']
    for x in range(len(coloum1)):
        if x>=6:
            a = coloum1[x].value
            # ans = naac(a)
            # data.append()
            doc = nlp(a)
            # # a.lower()
            # # list = a.split(",")
            # # list1 = list[0].split("[")
            # # list2 = list1[0].split("(")
            # # list3 = list2[0].split("\n")
            # # print(list3[0].lower())
            #
            tokens = [token.text for token in doc]
            # q = tokens[0].lower().strip(".',( ")
            # w = tokens[1].lower().strip(".',( ")
            # e = tokens[2].lower().strip(".',( ")
            # f = tokens[3].lower().strip(".',( ")
            lenn = len(tokens)
            ans = ""
            if lenn >= 3:
                q = tokens[0].lower().strip(".',( ")
                w = tokens[1].lower().strip(".',( ")
                e = tokens[2].lower().strip(".',( ")
                ans = q + w + e
            elif lenn == 2:
                q = tokens[0].lower().strip(".',( ")
                w = tokens[1].lower().strip(".',( ")
                ans = q + w
            else:
                q = tokens[0].lower().strip(".',( ")
                ans = q

            data.append(ans)
            # data.append(q+w+e)
    # print(data)

    # print(f"{x} {q} {w} {e}")

    cgpa = []
    coloum2 = ws['E']
    for y in range(len(coloum2)):
        if y>=6:
            b = str(coloum2[y].value)
            cgpa.append(b)
            # doc = nlp(b)
            # tokens = [token.text for token in doc]
            # print(tokens)
    # print("college CGPA:", cgpa)

    grade = []
    coloum3 = ws['F']
    for z in range(len(coloum3)):
        if z>=6:
            c = coloum3[z].value
            grade.append(c)
    # print("\n\nGrades:", grade)

    valid = []
    coloum4 = ws['G']
    for u in range(len(coloum4)):
        if u>=6:
            # d = str(coloum4[u].value)
            d = str(coloum4[u].value).rstrip('00:00:00')
            valid.append(d)
    # print("\n\nvalidation upto:", valid)

    # print(f"{cgpa[0]} {grade[0]} {valid[0]}")
    naac11 = []
    for j in range(len(data)):
        d1 = {"__id": j, "instituteName": data[j], "NAAC_grade": grade[j], "CGPA": cgpa[j],
              "valid  upto": valid[j]}
        naac11.append(d1)
    naac1 += naac11

# print(naac1)

# cc = '3.05'
# hello =[]
# for i in naac1:
#     if (i['CGPA'] == cc):
#         hello.append(i)
#
# print(hello)
# uni = "drIEMS"
# uni = uni.lower().strip("! @ #")
# print(uni)
# if(len(hello)>1):
#     for un  in hello:
#         if(un['instituteName'] == uni):
#             print(un)

# def compare_strings(a, b):
#     str1 = a.lower()
#     str2 = b.lower()
#     size = max(len(str1), len(str2))
#     count = 0
#
#     for i in range(size):
#         if str1[i] == str2[i]:
#             count += 1
#     if count == len(str1):
#         print("\n\nstring is same")
#     else:
#         print("\n\nkataiiii same nhi he")
#
# compare_strings("kjfefggg", "kshitiz")
with open("naac11.json", "w") as external_file:
    json.dump(naac1, external_file)
    external_file.close()