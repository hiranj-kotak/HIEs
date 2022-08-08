import spacy
from openpyxl import load_workbook
nlp = spacy.load("en_core_web_sm")
def checker(fp):
    file_path = fp
    wb = load_workbook(file_path)
    # sheet = ['sheet1']
    ws = wb.active
    institution_name = ws['A']
    # nba = []
    # for s in sheet:
    #     ws = wb[s]
    #     data = []
    #     coloumn1 = ws['A']

    UG1 = []
    name = []
    for x in range(len(institution_name)):
        if x>=2:
            a = institution_name[x].value
            doc1 = nlp(a)
            tokens = [token.text.lower().strip(".',()") for token in doc1]
            # print(tokens)
            lenn1 = len(tokens)
            ans1 = ""
            if lenn1 >= 3:
                q1 = tokens[0]
                w1 = tokens[1]
                e1 = tokens[2]
                ans1 = q1+w1+e1
            elif lenn1 >= 2:
                q1 = tokens[0]
                w1 = tokens[1]
                ans1 = q1+w1
            else:
                q1 = tokens[0]
                ans1 = q1
            name.append(ans1)
    # print("Name of Institute: ", name)

    pro = []
    program = ws['B']
    for y in range(len(program)):
        if y>=2:
            b = program[y].value
            try:
                doc2 = nlp(b)
            except:
                pass
            tokens = [token.text.lower().strip(" . ' , ( ) \n -").replace("engg", "engineering").replace("tech", "technology") for token in doc2]
            lenn2 = len(tokens)
            ans2 = ""
            if lenn2 >= 6:
                q2 = tokens[0]
                w2 = tokens[1]
                e2 = tokens[2]
                r2 = tokens[3]
                t2 = tokens[4]
                u2 = tokens[5]
                ans2 = q2+w2+e2+r2+t2+u2
            elif lenn2 >= 5:
                q2 = tokens[0]
                w2 = tokens[1]
                e2 = tokens[2]
                r2 = tokens[3]
                t2 = tokens[4]
                ans2 = q2 + w2 + e2 + r2 + t2
            elif lenn2 >= 4:
                q2 = tokens[0]
                w2 = tokens[1]
                e2 = tokens[2]
                r2 = tokens[3]
                ans2 = q2 + w2 + e2 + r2
            elif lenn2 >= 3:
                q2 = tokens[0]
                w2 = tokens[1]
                e2 = tokens[2]
                ans2 = q2 + w2 + e2
            elif lenn2 >= 2:
                q2 = tokens[0]
                w2 = tokens[1]
                ans2 = q2 + w2
            elif lenn2 >= 1:
                q2 = tokens[0]
                ans2 = q2
            else:
                try:
                    pass
                except:
                    print("A")

            pro.append(ans2)
    # print("program name: ", pro)

    stat = []
    status = ws['D']
    for z in range(len(status)):
        if z>=2:
            c = status[z].value
            # doc3 = nlp(c)
            # tokens = [token.text.lower() for token in doc3]
            stat.append(c)
    # print("Status: ", stat)
    nba1 = []
    for i in range(len(name)):
        d1 = {"__id": i+1, "INSname": name[i], "program": pro[i], "status": stat[i]}
        nba1.append(d1)
    UG1 += nba1
    print("\n\nENGINEERING UG TIER 1:\n", UG1)
    #         for i in tokens:
    #             # print(i)
    #             kk = i.lower().strip(" . ' , ( ) \n -").replace("engg", "engineering")
    #             # print(kk)
    #             pro.append(kk)
    # print(pro)
                # kk = tokens[i].lower().strip(".',()")
    # print("programs: ", pro)
# engg_ug_1()

# def engg_ug_2():
#     file_path = "engineering ug tier 2.xlsx"
#     wb = load_workbook(file_path)
#     ws = wb.active
#     institution_name = ws['A']
#     UG2 = []
#     name = []
#     for x in range(len(institution_name)):
#         if x>=2:
#             a = institution_name[x].value
#             doc1 = nlp(a)
#             tokens = [token.text.lower().strip(".',()") for token in doc1]
#             # print(tokens)
#             lenn1 = len(tokens)
#             ans1 = ""
#             if lenn1 >= 3:
#                 q1 = tokens[0]
#                 w1 = tokens[1]
#                 e1 = tokens[2]
#                 ans1 = q1 + w1 + e1
#             elif lenn1 >= 2:
#                 q1 = tokens[0]
#                 w1 = tokens[1]
#                 ans1 = q1 + w1
#             else:
#                 q1 = tokens[0]
#                 ans1 = q1
#             name.append(ans1)
#     # print("Name of Institute: ", name)
#
#     pro = []
#     program = ws['B']
#     for y in range(len(program)):
#         if y >= 2:
#             b = program[y].value
#             doc2 = nlp(b)
#             tokens = [
#                 token.text.lower().strip(" . ' , ( ) \n -").replace("engg", "engineering").replace("tech", "technology")
#                 for token in doc2]
#             lenn2 = len(tokens)
#             ans2 = ""
#             if lenn2 >= 6:
#                 q2 = tokens[0]
#                 w2 = tokens[1]
#                 e2 = tokens[2]
#                 r2 = tokens[3]
#                 t2 = tokens[4]
#                 u2 = tokens[5]
#                 ans2 = q2 + w2 + e2 + r2 + t2 + u2
#             elif lenn2 >= 5:
#                 q2 = tokens[0]
#                 w2 = tokens[1]
#                 e2 = tokens[2]
#                 r2 = tokens[3]
#                 t2 = tokens[4]
#                 ans2 = q2 + w2 + e2 + r2 + t2
#             elif lenn2 >= 4:
#                 q2 = tokens[0]
#                 w2 = tokens[1]
#                 e2 = tokens[2]
#                 r2 = tokens[3]
#                 ans2 = q2 + w2 + e2 + r2
#             elif lenn2 >= 3:
#                 q2 = tokens[0]
#                 w2 = tokens[1]
#                 e2 = tokens[2]
#                 ans2 = q2 + w2 + e2
#             elif lenn2 >= 2:
#                 q2 = tokens[0]
#                 w2 = tokens[1]
#                 ans2 = q2 + w2
#             else:
#                 q2 = tokens[0]
#                 ans2 = q2
#
#             pro.append(ans2)
#     # print("program name: ", pro)
#
#     stat = []
#     status = ws['D']
#     for z in range(len(status)):
#         if z >= 2:
#             c = status[z].value
#             doc3 = nlp(c)
#             tokens = [token.text.lower() for token in doc3]
#             stat.append(tokens)
#     # print("Status: ", stat)
#
#     nba2 = []
#     for i in range(len(name)):
#         d1 = {"__id": i, "INSname": name[i], "program": pro[i], "status": stat[i]}
#         nba2.append(d1)
#     UG2 += nba2
#     print("\n\nENGINEERING UG TIER 2:\n", UG2)
# # engg_ug_2()
#
# def engg_ug_3():
#     file_path = "engineering pg.xlsx"
#     wb = load_workbook(file_path)
#     ws = wb.active
#     institution_name = ws['A']
#     PG = []
#     name = []
#     for x in range(len(institution_name)):
#         if x>=2:
#             a = institution_name[x].value
#             doc1 = nlp(a)
#             tokens = [token.text.lower().strip(".',()") for token in doc1]
#             # print(tokens)
#             lenn1 = len(tokens)
#             ans1 = ""
#             if lenn1 >= 3:
#                 q1 = tokens[0]
#                 w1 = tokens[1]
#                 e1 = tokens[2]
#                 ans1 = q1 + w1 + e1
#             elif lenn1 >= 2:
#                 q1 = tokens[0]
#                 w1 = tokens[1]
#                 ans1 = q1 + w1
#             else:
#                 q1 = tokens[0]
#                 ans1 = q1
#             name.append(ans1)
#     # print("Name of Institute: ", name)
#
#     pro = []
#     program = ws['B']
#     for y in range(len(program)):
#         if y >= 2:
#             b = program[y].value
#             doc2 = nlp(b)
#             tokens = [
#                 token.text.lower().strip(" . ' , ( ) \n -").replace("engg", "engineering").replace("tech", "technology")
#                 for token in doc2]
#             lenn2 = len(tokens)
#             ans2 = ""
#             if lenn2 >= 6:
#                 q2 = tokens[0]
#                 w2 = tokens[1]
#                 e2 = tokens[2]
#                 r2 = tokens[3]
#                 t2 = tokens[4]
#                 u2 = tokens[5]
#                 ans2 = q2 + w2 + e2 + r2 + t2 + u2
#             elif lenn2 >= 5:
#                 q2 = tokens[0]
#                 w2 = tokens[1]
#                 e2 = tokens[2]
#                 r2 = tokens[3]
#                 t2 = tokens[4]
#                 ans2 = q2 + w2 + e2 + r2 + t2
#             elif lenn2 >= 4:
#                 q2 = tokens[0]
#                 w2 = tokens[1]
#                 e2 = tokens[2]
#                 r2 = tokens[3]
#                 ans2 = q2 + w2 + e2 + r2
#             elif lenn2 >= 3:
#                 q2 = tokens[0]
#                 w2 = tokens[1]
#                 e2 = tokens[2]
#                 ans2 = q2 + w2 + e2
#             elif lenn2 >= 2:
#                 q2 = tokens[0]
#                 w2 = tokens[1]
#                 ans2 = q2 + w2
#             else:
#                 q2 = tokens[0]
#                 ans2 = q2
#
#             pro.append(ans2)
#     # print("program name: ", pro)
#
#     stat = []
#     status = ws['D']
#     for z in range(len(status)):
#         if z >= 2:
#             c = status[z].value
#             doc3 = nlp(c)
#             tokens = [token.text.lower() for token in doc3]
#             stat.append(tokens)
#     # print("Status: ", stat)
#
#     nba2 = []
#     for i in range(len(name)):
#         d1 = {"__id": i, "INSname": name[i], "program": pro[i], "status": stat[i]}
#         nba2.append(d1)
#     PG += nba2
#     print("\n\nENGINEERING UG TIER 2:\n", PG)
# # engg_ug_3()

print("Enter the desired value:\nENGINEERING:\n\t1. Engineering UG tier 1\n\t2. Engineering UG tier 2\n\t3. Engineering PG\n\t4. Engineering Diploma")
print("MANAGEMENT:\n\t5.Management PG")
print("PHARMACY:\n\t6. Pharmacy UG\n\t7. Pharmacy Diploma")
print("8. MCA")
print("9. ARCHITECTURE")
print("10. HOSPTALITY")

x = input()
x = int(x)

if x == 1:
    file = "engineering ug tier 1.xlsx"
    checker(file)
elif x == 2:
    file = "engineering ug tier 2.xlsx"
    checker(file)
elif x == 3:
    file = "engineering pg.xlsx"
    checker(file)
elif x == 4:
    file = "engineering diploma.xlsx"
    checker(file)
elif x == 5:
    file = "management pg.xlsx"
    checker(file)
elif x == 6:
    file = "pharmacy ug.xlsx"
    checker(file)
elif x == 7:
    file = "pharmacy diploma.xlsx"
    checker(file)
elif x == 8:
    file = "mca.xlsx"
    checker(file)
elif x == 9:
    file = "architecture.xlsx"
    checker(file)
else:
    file = "hospitality and tourism management.xlsx"
    checker(file)
