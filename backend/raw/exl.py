from openpyxl import load_workbook
file_path = "Institutions_accredited_by_NAAC_whose_accreditation_period_is_valid_08_07_22.xlsx"
wb = load_workbook(file_path)
ws = wb["Universities"]


coloum = ws['C']
data = []
for  x in range(len(coloum)):
    a = coloum[x].value
    data.append(a)


coloum4 = ws['F']
data1 = []
for  y in range(len(coloum4)):
    b = coloum4[y].value
    data1.append(b)


for x in range(6):
    del data[0]
    del  data1[0]

d = dict(zip(data,data1))
print(d)

with open("naac.txt", "w") as external_file:
    print(d, file=external_file)
    external_file.close()
# for item in d:
#     if item == 'Rashtriya Sanskrit Vidyapeetha':
#         print(d[item])